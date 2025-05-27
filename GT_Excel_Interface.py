
from __future__ import division


from openpyxl import load_workbook
from openpyxl import Workbook

from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import time



class Test_File(object):  

    def __init__(self, choice, *argv):

        print(len(argv))
        if len(argv) != 0:
            self.filename = argv[0]

        if choice == "Read":
            self.wb = load_workbook(filename=self.filename, read_only = True, data_only = True)
            self.sheet_array = self.wb.sheetnames
        elif choice == "Write":
            self.wb = Workbook()     #  start new workbook
            self.results = self.wb.active      #  create new sheet
            self.sheet_array = self.wb.sheetnames

            self.active_row = 5
            self.i = 0 
            
        else:
            self.wb = load_workbook(filename=self.filename, read_only = False, data_only = True)
            self.sheet_array = self.wb.sheetnames 
            self.name = self.filename
        


    '''
    Generic Methods For Reading
    =======================================================================================
    read_cell(self, my_col, my_row, tab) Reads a specific cell
        Inputs
            my_col(Int) - what col the cell is at
            my_row(Int) - what row the cell is at 
            Tab(Int) - what tab we want read from
        Changes
            N/A
        Outputs
            value(Str) - value in the cell
        

    read_row(self, tab, my_col, my row) Reads a line unit it hits an empty value
        Inputs
            tab(Int) - what tab in the excel are we reading from
            my_col(Int) - what column we want to start reading from
            my_row(Int) - what row we want to read
        Changes
            N/A
        Outputs
            values(Array) - All values in the cell 
        
    read_row_for_length(self, tab, my_col, my row, len) Reads a line for a specified length
        Inputs
            sheet(int) - what tab in the excel are we reading from
            my_col(Int) - what column we want to start reading from
            my_row(Int) - what row we want to read
        Changes
            N/A
        Outputs
            values(Array) - All values in the cell 

    read_column_for_length(self, tab, my_col, my row, len) Reads a column until empty
        Inputs
            sheet(int) - what tab in the excel are we reading from
            my_col(Int) - what column we want to start reading from
            my_row(Int) - what row we want to read
        Changes
            N/A
        Outputs
            values(Array) - All values in the cell

    read_column_for_length(self, tab, my_col, my row, len) Reads a column for a specified length
        Inputs
            sheet(int) - what tab in the excel are we reading from
            my_col(Int) - what column we want to start reading from
            my_row(Int) - what row we want to read
        Changes
            N/A
        Outputs
            values(Array) - All values in the cell 


    def close_file(self) Closes the currently open excel file
        Inputs
            N/A
        Changes
            wb(excel file) 
        Outpus
            N/A
    '''
  
    def read_cell(self, my_col, my_row, tab):
        ws = self.wb[self.sheet_array[tab]]
        value = ws.cell(column = my_col, row = my_row).value

        return value

    
    def read_row(self, my_col, my_row, tab):
        values = []
        ws = self.wb[self.sheet_array[tab]]
        while True:
            val = ws.cell(column = my_col, row = my_row).value
            if val == None:
                break
            
            values.append(val)  #Add the excel data an array 
            my_col = my_col + 1 #Move on to next Col 

        return values #Return array with all the values 


    def read_row_for_length(self, my_col, my_row, tab, my_len):
        values = []
        ws = self.wb[self.sheet_array[tab]]
        for i in range(my_len):
            val = ws.cell(column = my_col, row = my_row).value
            if val == None:
                break
            
            values.append(val)  #Add the excel data an array 
            my_col = my_col + 1 #Move on to next Col 

        return values #Return array with all the values
    


    def read_comlumn(self, my_col, my_row, tab):
        values = []
        ws = self.wb[self.sheet_array[tab]]
        while True:
            val = ws.cell(column = my_col, row = my_row).value
            if val == None:
                break
            
            values.append(val)  #Add the excel data an array 
            my_row = my_row + 1 #Move on to next Col 



    def read_column_for_length(self, my_col, my_row, tab, my_len):
        values = []
        ws = self.wb[self.sheet_array[tab]]
        for i in range(my_len):
            val = ws.cell(column = my_col, row = my_row).value
            if val == None:
                break
            
            values.append(val)  #Add the excel data an array 
            my_row= my_row + 1 #Move on to next Row 

        return values #Return array with all the values
        
    def close_file(self):

        self.wb._archive.close()        
        self.wb.close()

    '''
    Specialized Methods For Reading
    =======================================================================================
    def read_string_from_cell(self, my_col, my_row, tab) Reads a string and strips out any blank spaces
        Inputs
            my_col(Int) - what col the cell is at
            my_row(Int) - what row the cell is at 
            Tab(Int) - what tab we want read from
        Changes
            N/A
        Outputs
            value(Str) - value in the cell

    '''
    
    def read_string_from_cell(self, my_col, my_row, tab):
        val = self.read_cell(my_col, my_row, tab)
        if val != None:
            val = val.strip()

        return val



    def read_main_settings(self, repos):

        ws = self.sheet_array[0]
        
        #Gets and translates the Frame
        frame = ws.cell(column = 3, row = 4).value


        #If the programs detects a new breaker setting, it sets a flag to recalibrate
        if repos.rating != int(ws.cell(column = 3, row = 5).value):
            repos.re_cal_needed = True

        #Gets the frame rating
        repos.rating = int(ws.cell(column = 3, row = 5).value)

        #Gets the standard and translates it
        standard = ws.cell(column = 3, row = 6).value
        if standard == "IEC":
            repos.breaker_protection_capacity['standard'] = 1
            repos.standard = 1
        elif standard == "Global":
            repos.breaker_protection_capacity['standard'] = 3
            repos.standard = 3

        repos.override_level = int(ws.cell(column = 3, row = 7).value)
        repos.configuration["Override"] = int(ws.cell(column = 3, row = 7).value)


        #Gets the information that will be displayed on the main tab of the result excel file 
        j = 3
        while True:

            val = ws.cell(column = 5, row = j).value
            
            if val != None and val != '':
                val = val.strip()
                repos.main_keys.append(val)
            else:
                break
            
            j = j+1


        




    '''
    Naming Methods 
    =======================================================================================

    def name_file(name): #Names file base on name 
        Inputs
            name(String) - String to be used as the file name
        Changes
            self.name(String) - The file name of this object 
        Outputs
            N/A

    def name_file_std_method(in_file, file_path, test_num) Names file based on proceedurally generated method
        Inputs
            in_file(String) - The name of the input file this report is being generated from
            file_path(String) - The directory that we want to save this file in
            test_num(Int) - What run we are on 
        Changes
            self.name(String) - The file name of this object 
        Outputs
    '''

    
    def name_file(self, name):
        self.name = name
        self.wb.save(self.name)

    def name_file_std_method(self, in_file, file_path, test_num):

        self.name = file_path + "\\" + "Result"  + '_' + str(test_num) + '_of_' + in_file              #  generate test report name
        print("\n", self.name, "\n")          

    '''
    Generic Methods For Writing
    =======================================================================================
     def create_tab(self, name): Creates a new tab in the excel file 

        self.new_excel.create_sheet(name)
        self.sheet_array = self.wb.sheetnames

     def write_cell(self, my_col, my_row, tab, val): Writes value to the specified cell on the excel sheet
        Inputs
            my_col(int) - column to be written
            my_row(int) - row to be written
            tab(int)    - tab to be written on 
        Changes
            Writes values to the specified cells 
        Outputs
            N/A
        
     def write_row_with_dictionary(self, my_col, my_row, tab, keys, dictionary): Writes a row of values 
        Inputs
            my_col(int) - starting column to be written
            my_row(int) - starting row to be written
            tab(int)    - tab to be written on
            key(array)  - array of the dictionary keys to be iterrated through
            dictionary (dictionary) - Dictionary with the values that will be written to the excel 
        Changes
            Writes values to the specified cells 
        Outputs
            N/A

     def write_column_with_dictionary(self, my_col, my_row, tab, keys, dictionary):
            my_col(int) - starting column to be written
            my_row(int) - starting row to be written
            tab(int)    - tab to be written on
            key(array)  - array of the dictionary keys to be iterrated through
            dictionary (dictionary) - Dictionary with the values that will be written to the excel 
        Changes
            Writes values to the specified cells 
        Outputs
            N/A

     def write_column_with_array(self, my_col, my_row, tab, my_array):
         Inputs
            my_col(int) - starting column to be written
            my_row(int) - starting row to be written
            tab(int)    - tab to be written on
            key(array)  - array of values to be written to the excel 
        Changes
            Writes values to the specified cells 
        Outputs
            N/A    

     def write_row_with_array(self, my_col, my_row, tab, my_array):
         Inputs
            my_col(int) - starting column to be written
            my_row(int) - starting row to be written
            tab(int)    - tab to be written on
            key(array)  - array of values to be written to the excel 
        Changes
            Writes values to the specified cells 
        Outputs
            N/A    

     def save_file(self):
        self.wb.save(self.Name)

        
     def close(self): Closes the file 
        self.results.close()
    '''

    def create_tab(self, name):

        self.wb.create_sheet(name)
        self.sheet_array = self.wb.sheetnames

    def write_cell(self, my_col, my_row, tab, val):
        
        ws = self.wb[self.sheet_array[tab]]
        ws.cell(column = my_col, row = my_row, value = val)
        
    def read_cell(self, my_col, my_row, tab):
        ws = self.wb[self.sheet_array[tab]]
        value = ws.cell(column = my_col, row = my_row).value

        return value

    def read_cell(self, my_col, my_row, tab):
        ws = self.wb[self.sheet_array[tab]]
        value = ws.cell(column = my_col, row = my_row).value

        return value
    
    def read_string_from_cell(self, my_col, my_row, tab):
        ws = self.wb[self.sheet_array[tab]]
        value = ws.cell(column = my_col, row = my_row).value
        value = str(value)
        if value != None:
            value = value.strip()

        return value
    
    def write_row_with_dictionary(self, my_col, my_row, tab, keys, dictionary):

        ws = self.wb[self.sheet_array[tab]]
        for key in keys:
            ws.cell(column = my_col, row = my_row, value = dictionary[key])
            my_col = my_col+ 1

    def write_column_with_dictionary(self, my_col, my_row, tab, keys, dictionary):

        ws = self.wb[self.sheet_array[tab]]
        for val in my_array:
            ws.cell(column = my_col, row = my_row, value = dictionary[key])
            my_row= my_row + 1

    def write_column_with_array(self, my_col, my_row, tab, my_array):

        ws = self.wb[self.sheet_array[tab]]
        for val in my_array: 
            ws.cell(column = my_col, row = my_row, value = val)
            my_row= my_row + 1

    def write_row_with_array(self, my_col, my_row, tab, my_array):

        ws = self.wb[self.sheet_array[tab]]
        for val in my_array: 
            ws.cell(column = my_col, row = my_row, value = val)
            my_col = my_col + 1       


    def save_file(self):
        self.wb.save(self.name)

        
    def close(self):
        self.wb.close()

    def get_tabs(self):
        self.sheet_array = self.wb.sheetnames


    '''
    Specialized Methods For Writing
    =======================================================================================
    def header(self) Creates header for the reports of the automated tests
        Inputs
            N/A
        Changes
            Writes values to certain cells on the test report
        Outputs
            N/A

    '''        
              
    def header(self):               #  writes header for test report
        A = [ 'Title', 'Tester:', 'Date']
        
        for i in range(0, len(A)):
            self.results.cell(column=1, row=i+1, value= A[i])
            i+=1
        
        B = ['General Test','David Arisumi ', time.asctime()]
        
        for i in range(0, len(B)):
            self.results.cell(column=2, row=i+1, value= B[i])
            i+=1
        self.wb.save(self.name)


  
        

            



