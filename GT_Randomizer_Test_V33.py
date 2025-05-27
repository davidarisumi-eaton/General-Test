from __future__ import division
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook

from openpyxl.styles.borders import Border, Side

import win32com
from win32com.client import Dispatch

import os

import time

import GT_Input_Reader

import random

def main():

    name = "SDPU"
    in_file = "Tests\\Generic\\" + name + ".xlsx"

    wb = load_workbook(filename=in_file)
    sheet_array = wb.sheetnames

    for k in range(5, 200):

        row = k
        col = 1
        ws = wb[sheet_array[3]]
        sp_array = create_setpoints()
        for val in sp_array:
            ws.cell(column = col, row = row, value = val)
            col = col+ 1

        ws = wb[sheet_array[1]]

        if sp_array[15] == 0:
            i_min = int(sp_array[17] * sp_array[7])
        else:
            i_min = int(sp_array[7])

        i_max = int(sp_array[7]*12)
        print(i_min)
        print(i_max)
        current = random.randint(i_min, i_max)

        ws.cell(column = 2, row = row, value = current)
        row = row 


    dir_path = str(os.getcwd()) 
    Title = dir_path + "\\Tests\\Created\\Random_Vals.xlsx"
    wb.save(Title)
    wb.close()
    
    time.sleep(4)


    #Opens and saves the xcel documents to change the formulas to actual values
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(Title)
    xlBook.Save()
    xlBook.Close()

            

        
    
    

def create_setpoints():

    rating = 125
    style_one = 0
    style_two = 0
    frame = 2
    
    ZSI_Array = [1,0]
    LD_Slope_Array = [2,3]
    LD_PU_Array = [45, 50, 63, 70, 80, 90, 100, 110, 125]
    LD_Time_Array = [.5, 1, 2, 4, 5, 6, 8, 10, 12, 15, 20, 24]
    SPU_Array = [1.5, 2, 3, 4, 5, 6, 8, 10, 12]
    SD_Slope_Array = [0, 1]
    SD_Time_Array = [.05, .1, .15, .2, .3, .5, .067]
    Inst_PU_Array = [2, 3, 4, 5, 6, 8, 10, 15, 20, 24]
    GF_PU_Array = [.2, .3, .4, .6, .8, 1, .5]
    GF_Time_Array = [.1, .15, .2, .3, .5, .76, 1]
    Neutral_Ratio_Array = [100, 60, 0]


    ld_thermal = 0
    zsi = random.choice(ZSI_Array)
    ldpu = random.randint(45, 125)
    lds = random.choice(LD_Slope_Array)
    
    if lds == 2:
        ld_time = random.randint(5,240)
        ld_time = ld_time/10
        
    else: 
        ld_time = random.randint(5,70)
        ld_time = ld_time/10

    hla_one = random.randint(50,119)
    sds = random.randint(0,1)

    if lds == 3:
        sds = 0
    
    sdpu = random.randint(15, 120)
    sdpu = sdpu/10
    
    if sds == 0:
        sdt = random.randint(5, 50)
        sdt = sdt/100
    else:
        sdt = random.randint(6, 50)
        if sdt == 6:
            sdt = 6.7
            
        sdt = sdt/100
        
    inst = random.randint(2, 12)

    gf_type = 0
    gf_mode = random.randint(0,2)

    gfs = random.randint(0,1)

    if lds == 3 or sds == 1:
        gfs = 0

    gfpu = random.randint(20,100)
    gfpu = gfpu/100

    if gfs == 0:
        gft = random.randint(5, 100)
        gft = gft/100
    else:
        gft = random.randint(67, 100)
        gft = gft/100

    gf_therm = 0
    neut = random.choice(Neutral_Ratio_Array)
    hla_two = random.randint(hla_one+1, 120)
    gf_alm = random.randint(50,100)
    gf_alm = round(gf_alm/5)*5
    thm_alm = 85
    
    sp_array = [rating,
             frame,
             style_one,
             style_two,
             ld_thermal,
             zsi,
             lds,
             ldpu, 
             ld_time,
             hla_one,
             sds,
             sdpu,
             sdt,
             inst,
             gf_type,
             gf_mode,
             gfs,
             gfpu,
             gft,
             gf_therm,
             neut,
             hla_two,
             gf_alm,
             thm_alm]

    return sp_array
    
main()
